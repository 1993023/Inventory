
# coding: utf-8

# In[111]:

import pandas as pd
import numpy as np
import os

file_path = os.path.dirname(os.path.abspath( __file__ ))

# In[153]:

def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False, 
                       **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    Parameters:
      filename : File path or existing ExcelWriter
                 (Example: '/path/to/file.xlsx')
      df : dataframe to save to workbook
      sheet_name : Name of sheet which will contain DataFrame.
                   (default: 'Sheet1')
      startrow : upper left cell row to dump data frame.
                 Per default (startrow=None) calculate the last row
                 in the existing DF and write to the next row...
      truncate_sheet : truncate (remove and recreate) [sheet_name]
                       before writing DataFrame to Excel file
      to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                        [can be dictionary]

    Returns: None
    """
    from openpyxl import load_workbook

    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl')

    # Python 2.x: define [FileNotFoundError] exception if it doesn't exist 
    try:
        FileNotFoundError
    except NameError:
        FileNotFoundError = IOError


    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)#,read_only=False , keep_vba=True)

        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)

        # copy existing sheets
        writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()


############## Condition 1 ###############

def leadTime(val1,val2):
    lead_time = None
    s6 = ex.parse('Sheet6')
    lead_time = s6.loc[(s6['Plant']==val1),val2]
    if len(lead_time) > 0:
        lead_time = s6.loc[(s6['Plant']==val1),val2].values[0]
    return lead_time

def similarity_index(val1,val2):
    sim_score = None
    s7 = ex.parse('Sheet7')
    sim_score = s7.loc[(s7['Wheat']==val1),val2]
    if len(sim_score)>0:
        sim_score = s7.loc[(s7['Wheat']==val1),val2].values[0]
    return sim_score


def threshold_check(sim_score,under_supp):
    if sim_score >= 0.75 and under_supp == 'N':
        return 'Ok'
    else:
        return 'Not ok'

def underSupply(val1,val2):
    under_supp =None
    s5 = ex.parse('Sheet5')
    under_supp = s5.loc[(s5['Plant']==val1),val2]
    if len(under_supp)>0:
        under_supp = s5.loc[(s5['Plant']==val1),val2].values[0]
    return under_supp

def replenishment(val1,val2):
    replen =None
    s3 = ex.parse('Sheet3')
    replen = s3.loc[(s3['Plant']==val1),val2]
    if len(replen)>0:
        replen = s3.loc[(s3['Plant']==val1),val2].values[0]
    return replen

def daysOnHand(val1,val2):
    days_on =None
    s4 = ex.parse('Sheet4')
    days_on = s4.loc[(s4['Plant']==val1),val2]
    if len(days_on)>0:
        days_on = s4.loc[(s4['Plant']==val1),val2].values[0]
    return days_on


def updateValExcel(sheet_name,df,index_set,index_r,col,value):
    df.set_index(index_set, inplace=True)
    df[col][index_r] = value
    df.reset_index(level=0, inplace=True)
    updatedftoexcel(df,sheet_name)
    
def updatedftoexcel(df,sheetName):
    append_df_to_excel(file_path+'/Grist Optimizer changed.xlsx', df, sheet_name=sheetName,startrow=0,index=False)           
    

def withInPlant(plant,ptype,d_val):
    ex = pd.ExcelFile(file_path+'/Grist Optimizer changed.xlsx')
    s2 = ex.parse('Sheet2')
    temp = pd.DataFrame() 
    for i in p_type :
        temp = temp.append(s2.loc[(s2['Plant']==plant) & (s2[i]>d_val),i])
    
    #temp = temp.reset_index(drop =True)
    ## Else to check for other plants as well
    print('Entered withInPlant function')

    if temp.empty == False:
        print('Found some space in same plant lets check for Under suppy---')
        temp.reset_index(level=0, inplace=True)
        temp.columns = ['Type','Values']
        temp = temp.sort_values(by=['Values'],ascending=False)
        temp = temp.loc[(temp['Values']>0)]
        temp = temp.reset_index(drop =True)

## Only value then no need to check for Similarity index
#         if len(temp)==1:
#             print('Only one row')
#             temp.reset_index(drop=True)
#             return temp
#         else :

#         if len(temp):
        
        temp = temp.reset_index(drop=True)
        list_a = []
        for j in range(len(temp)):
            sim_score = similarity_index(temp['Type'][j],ptype)
            if sim_score:
                list_a.append(sim_score)
            else :
                list_a.append('')
        temp['Similarity Score'] = list_a
        temp['Plant'] = plant
        temp = temp.sort_values(by='Similarity Score',ascending=False)
        temp = temp.reset_index(drop=True)

        list_u = []
        for j in range(len(temp)):
            under_supp = underSupply(plant,temp['Type'][j])
            if under_supp:
                list_u.append(under_supp)
            else:
                list_u.append('')
        temp['Under Supply'] = list_u
        temp = temp.sort_values(by='Under Supply',ascending=True)
        temp = temp.reset_index(drop=True)            

        list_t =[]
        for j in range(len(temp)):
            thresh_check = threshold_check(temp['Similarity Score'][j],temp['Under Supply'][j])
            list_t.append(thresh_check)

        temp['Threshold Check'] = list_t
        temp = temp.sort_values(by='Threshold Check',ascending=False)
        temp = temp.reset_index(drop=True)            

        print('temp---------',temp)

        ## Check for ok condition from available rows and pick one
        temp_r1 = None
        for i in range(len(temp)):
            if temp['Threshold Check'][i] == 'Ok':
                ## Updating(subtracting values) value after demand satisfied.
#                 updateValExcel('Sheet2',s2,'Plant',temp['Plant'][i],temp['Type'][i],  temp['Values'][i]-d_val)
                temp_r1 = 'Please pick the value from Unit - ('+str(temp['Plant'][i])+','+str(temp['Type'][i])+') with current quantity '+str(temp['Values'][i])
                #temp_r1 = temp[i:i+1].to_dict()
                break
        print('temp_r1',temp_r1)
        if temp_r1 :
            return temp_r1#[0:1]
        else :
            ##Call function for condition 2
            print('Check for Condition 2')
            return otherPlant(plant,ptype,d_val)
            
    else:
        print('Check for condition 2')
        return otherPlant(plant,ptype,d_val)
#         return None



# temp = withInPlant('PH','Low Protein',60000.0)
# temp = withInPlant('TCB','Canadian',100000.0)
# temp

# similarity_index('Canadian','Canadian')


# In[177]:


def demand_fill(plant,ptype,d_val):
    
    #temp = withInPlant('TCB','Canadian',50000.0)
    temp = withInPlant(plant,ptype,d_val)
    if temp:
        return temp
    else:
        return None

def demandPriority(val1,val2):
    pri_score = 0
    s9 = ex.parse('Sheet9')
    pri_score = s9.loc[(s9['Plant']==val1),val2]
    if len(pri_score) > 0:
        pri_score = s9.loc[(s9['Plant']==val1),val2].values[0]
    return pri_score



# In[178]:

## Condition2

def condition2(df,ptype,d_val):
    print('Entered in Condition2')
    df = df.reset_index(drop=True)
    df= pd.melt(df,id_vars=["Plant", "Category"],var_name="Type",value_name="Values")
    df = df.drop_duplicates()
    temp = pd.DataFrame()
    for j in range(len(df)):
        if df['Values'][j] > d_val:
            temp = temp.append([{'Type':df['Type'][j],'Plant':df['Plant'][j],'Category':df['Category'][j],'Values':df['Values'][j]}],ignore_index=True)
        
    temp = temp.sort_values(by=['Values'],ascending=False)
    temp = temp.reset_index(drop=True)
    return temp
    
    

def otherPlant(plant,ptype,d_val):
    print('Entered in otherPlant')
    ex = pd.ExcelFile(file_path + '/Grist Optimizer changed.xlsx')
    s2 = ex.parse('Sheet2')
    temp1 = pd.DataFrame() 
    for i in p_type :
        temp1 = temp1.append(s2.loc[(s2[i]>d_val),['Plant','Category','Canadian','Lithuanian','Russian','German','French',
                                                   'Argentinaian','Ukrainian','SRW - US']])
    ## Else to check for other plants as well
    if temp1.empty == False:
        temp1 = temp1.drop_duplicates()
        temp = condition2(temp1,ptype,d_val)
#         print(temp)

        list_a = []
        for j in range(len(temp)):
            sim_score = similarity_index(temp['Type'][j],ptype)
            if sim_score:
                list_a.append(sim_score)
            else :
                list_a.append('')
        temp['Similarity Score'] = list_a
        temp = temp.sort_values(by='Similarity Score',ascending=False)
        temp = temp.reset_index(drop=True)

        list_u = []
        for j in range(len(temp)):
            under_supp = underSupply(temp['Plant'][j],temp['Type'][j])
            if under_supp:
                list_u.append(under_supp)
            else:
                list_u.append('')
        temp['Under Supply'] = list_u
        temp = temp.sort_values(by='Under Supply',ascending=True)
        temp = temp.reset_index(drop=True)

        list_l = []
        for j in range(len(temp)):
            lead_time = leadTime(plant,temp['Plant'][j])
            if lead_time is not None:
                list_l.append(lead_time)
            else:
                list_l.append('')
        temp['Lead Time'] = list_l
        temp = temp.sort_values(by='Lead Time',ascending=True)
        temp = temp.reset_index(drop=True)
        
        list_t =[]
        for j in range(len(temp)):
            thresh_check = threshold_check(temp['Similarity Score'][j],temp['Under Supply'][j])
            list_t.append(thresh_check)

        temp['Threshold Check'] = list_t
        temp = temp.sort_values(by='Threshold Check',ascending=False)
        temp = temp.reset_index(drop=True)  

        ########## Sorting again
        temp = temp[temp['Threshold Check'] != 'Not ok']
        temp = temp.sort_values(by='Threshold Check',ascending=False)
        temp = temp.sort_values(by='Lead Time',ascending=True)
        temp = temp.reset_index(drop=True)

        print(temp)
        ## Check for ok condition from available rows and pick one
        temp_r1 = None
        for i in range(len(temp)):
            if temp['Threshold Check'][i] == 'Ok':
                ## Updating(subtracting values) value after demand satisfied.
#                 updateValExcel('Sheet2',s2,'Plant',temp['Plant'][i],temp['Type'][i],  temp['Values'][i]-d_val)
                temp_r1 = 'Please pick the value from Unit - ('+str(temp['Plant'][i])+','+str(temp['Type'][i])+') with current quantity '+str(temp['Values'][i])
                #temp_r1 = temp[i:i+1].to_dict()
                break
        
        if temp_r1 :
            print('return',temp_r1)
            return temp_r1#[0:1]
        else :
            ##Call function for condition 2
            print('Check for Condition 3')
            
            return multiplePlant(plant,ptype,d_val)
            
    else:
        print('Check for condition 3')
        return multiplePlant(plant,ptype,d_val)




# In[ ]:




# In[180]:

## Condition 3
def condition3(df,ptype,d_val):
    print('Entered in Condition3')
    df = df.reset_index(drop=True)
    df= pd.melt(df,id_vars=["Plant", "Category"],var_name="Type",value_name="Values")
    df = df.drop_duplicates()
    temp = pd.DataFrame()
    for j in range(len(df)):
        if df['Values'][j] > d_val/2:
            temp = temp.append([{'Type':df['Type'][j],'Plant':df['Plant'][j],'Category':df['Category'][j],'Values':df['Values'][j]}],ignore_index=True)
        
    temp = temp.sort_values(by=['Values'],ascending=False)
    temp = temp.reset_index(drop=True)
    return temp

##    
   

def multiplePlant(plant,ptype,d_val):
    print('Entered in multiplePlant')
    ex = pd.ExcelFile(file_path + '/Grist Optimizer changed.xlsx')
    s2 = ex.parse('Sheet2')
    temp1 = pd.DataFrame() 
    for i in p_type :
        temp1 = temp1.append(s2.loc[(s2[i]>d_val/2),['Plant','Category','Canadian','Lithuanian','Russian','German','French',
                                                   'Argentinaian','Ukrainian','SRW - US']])
    ## Else to check for other plants as well
    if temp1.empty == False:
        temp1 = temp1.drop_duplicates()
        temp = condition3(temp1,ptype,d_val)
#         print(temp)

        list_a = []
        for j in range(len(temp)):
            sim_score = similarity_index(temp['Type'][j],ptype)
            if sim_score:
                list_a.append(sim_score)
            else :
                list_a.append('')
        temp['Similarity Score'] = list_a
        temp = temp.sort_values(by='Similarity Score',ascending=False)
        temp = temp.reset_index(drop=True)

        list_u = []
        for j in range(len(temp)):
            under_supp = underSupply(temp['Plant'][j],temp['Type'][j])
            if under_supp:
                list_u.append(under_supp)
            else:
                list_u.append('')
        temp['Under Supply'] = list_u
        temp = temp.sort_values(by='Under Supply',ascending=True)
        temp = temp.reset_index(drop=True)

        list_l = []
        for j in range(len(temp)):
            lead_time = leadTime(plant,temp['Plant'][j])
            if lead_time is not None:
                list_l.append(lead_time)
            else:
                list_l.append('')
        temp['Lead Time'] = list_l
        temp = temp.sort_values(by='Lead Time',ascending=True)
        temp = temp.reset_index(drop=True)
        
        list_t =[]
        for j in range(len(temp)):
            thresh_check = threshold_check(temp['Similarity Score'][j],temp['Under Supply'][j])
            list_t.append(thresh_check)

        temp['Threshold Check'] = list_t
        temp = temp.sort_values(by='Threshold Check',ascending=False)
        temp = temp.reset_index(drop=True)  
###
        list_r = []
        for j in range(len(temp)):
            replen = replenishment(plant,temp['Type'][j])
            if replen:
                list_r.append(replen)
            else:
                list_r.append('')
        temp['Replenishment'] = list_r
#         temp = temp.sort_values(by='Replenishment',ascending=True)
        temp = temp.reset_index(drop=True)

        list_d = []
        for j in range(len(temp)):
            days_on = daysOnHand(plant,temp['Type'][j])
            if days_on:
                list_d.append(days_on)
            else:
                list_d.append('')
        temp['Days On Hand'] = list_d
#         temp = temp.sort_values(by='Days On Hand',ascending=True)
        temp = temp.reset_index(drop=True)

        temp['DaysOnHand > Replenishment'] = np.where(temp['Days On Hand'] > temp['Replenishment'],'Yes','No')
        temp = temp.sort_values(by='DaysOnHand > Replenishment',ascending=False)

        ########## Sorting again
        temp = temp[temp['Threshold Check'] != 'Not ok']
        temp = temp[temp['DaysOnHand > Replenishment'] != 'No']
        temp = temp.sort_values(by='Threshold Check',ascending=False)
        temp = temp.sort_values(by='Lead Time',ascending=True)
        temp = temp.reset_index(drop=True)


        print(temp)                                          
        ## Check for ok condition from available rows and pick one
        temp_r1 = None 
        break_temp = 0
        ## check condition to pick output
        for i in range(len(temp)):
            for j in range(1,len(temp)):
                if temp['Threshold Check'][i] == 'Ok' and temp['DaysOnHand > Replenishment'][i]=='Yes' and temp['Threshold Check'][j] == 'Ok' and temp['DaysOnHand > Replenishment'][j]:
                    if (temp['Values'][i]*0.9 + temp['Values'][j]*0.9) >= d_val:
                        
                        rest = d_val - temp['Values'][i] * 0.9
                        print('Plant1 : ',temp['Values'][i]-temp['Values'][i]*0.9,'\n','Plant2 : ',temp['Values'][j]-rest)
    #                     ##Updating(subtracting values) value after demand satisfied.
#                         updateValExcel('Sheet2',s2,'Plant',temp['Plant'][i],temp['Type'][i], temp['Values'][i]-temp['Values'][i]*0.9)
#                         updateValExcel('Sheet2',s2,'Plant',temp['Plant'][i+1],temp['Type'][i+1], temp['Values'][j]-rest)
                        temp_r1 = 'Please pick the value from Unit - ('+str(temp['Plant'][i])+','+str(temp['Type'][i])+' and '+str(temp['Plant'][j])+','+str(temp['Type'][j])+') with current quantity '+str(temp['Values'][i])+','+str(temp['Values'][j])+' respectively'
                        #temp_r1 = str(temp[i:i+1].to_dict()) +'\n' +str(temp[j:j+1].to_dict())
                        break_temp = 1
                        break
            if break_temp == 1:
                break
        
        if temp_r1 :
            print('return',temp_r1)
            return temp_r1
        else :
            print('No Success')
            return None
            
    else:
        print('No Success')
        return None


# In[181]:

#updateValExcel(sheet_name,df,index_set,index_r,col,value)
def edOrder(df,ptype) :
    ed_order = pd.DataFrame()
    for i in range(len(df)):
        for j in p_type:
            if df[j][i] > 0 :
                #print('hi')
                #print(df['Plant'][i],'---',j,'---',df[j][i],demandPriority(df['Plant'][i],j))
                pri_score = demandPriority(df['Plant'][i],j)

    #             list_ed.append(demandPriority(ed['Plant'][i],j))
                ed_order = ed_order.append([{'Category':df['Category'][i],'Plant':df['Plant'][i],'Type':j,'Value':df[j][i],'Demand Priority':pri_score}],ignore_index=True)

        #temp = demand_fill(ed['Plant'][i],j,ed[j][i])
    #             ed['Suggestion'][i] = temp
    #             updatedftoexcel(ed,'Emergency Demand (Tonnes)')
    if ed_order.empty == 'True':
        return None
    else :
        print('ed_order---','\n',ed_order)
        ed_order = ed_order.sort_values(by='Demand Priority',ascending=False)
        ed_order = ed_order.reset_index(drop=True)
        return ed_order 


ex = pd.ExcelFile(file_path+'/Grist Optimizer changed.xlsx')
ed = ex.parse('Emergency Demand (Tonnes)')
p_type = ['Canadian','SRW - US','Lithuanian','Russian','German','French','Argentinaian','Ukrainian']

## Removing all the previous values
for i in p_type:
    out = 'O-'+str(i)
    ed[out] = 0
   
ed_ordered = edOrder(ed,p_type)
ed_ordered['Suggestion'] = None

#Ordered demand calculation
for i in range(len(ed_ordered)):
    print(ed_ordered['Plant'][i],ed_ordered['Type'][i],ed_ordered['Value'][i])
    temp = demand_fill(ed_ordered['Plant'][i],ed_ordered['Type'][i],ed_ordered['Value'][i])
    print('###########################')
    ed_ordered['Suggestion'][i] = str(temp)
print('-----------------------------------------') 
#updating output to excel
for i in range(len(ed_ordered)) :
    print('Emergency Demand (Tonnes)','Plant',ed_ordered['Plant'][i],'O-'+ed_ordered['Type'][i],ed_ordered['Suggestion'][i])
    updateValExcel('Emergency Demand (Tonnes) - C',ed,'Plant',ed_ordered['Plant'][i],'O-'+ed_ordered['Type'][i],ed_ordered['Suggestion'][i])
    

## Refresh file 
# from win32com.client import Dispatch
# xl = Dispatch('Excel.Application')
# wb = xl.Workbooks.Open(file_path+'/Grist Optimizer changed.xlsx')
# # do some stuff
# wb.Close(True) # save the workbook

# import time
# time.sleep(2)

