import pandas as pd
import numpy as np
import math
import os
import datetime

file_path = os.path.dirname(os.path.abspath( __file__ ))


############################## Writing in excel 
##############################

def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False, 
                       **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    Parameters:
      filename : File path or existing ExcelWriter
                 (Example: '/path/to/file.xlsx')
      df : dataframe to save to workbook
      sheet_name : Name of sheet which will contain DataFrame.
                   (default: 'Sheet1')
      startrow : upper left cell row to dump data frame.
                 Per default (startrow=None) calculate the last row
                 in the existing DF and write to the next row...
      truncate_sheet : truncate (remove and recreate) [sheet_name]
                       before writing DataFrame to Excel file
      to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                        [can be dictionary]

    Returns: None
    """
    from openpyxl import load_workbook

    import pandas as pd

    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl')

    # Python 2.x: define [FileNotFoundError] exception if it doesn't exist 
    try:
        FileNotFoundError
    except NameError:
        FileNotFoundError = IOError


    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)

        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)

        # copy existing sheets
        writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()
    
    
def updateValExcel(sheet_name,df,index_set,index_r,col,value):
    df.set_index(index_set, inplace=True)
    df[col][index_r] = value
    df.reset_index(level=0, inplace=True)
    updatedftoexcel(df,sheet_name)
    
def updatedftoexcel(df,sheetName):
    append_df_to_excel(file_path + '/Inventory optimization working.xlsx', df, sheet_name=sheetName,startrow=0,index=False)   



def updateValExcel1(sheet_name,df,index_set,index_r,col,value):
    df.set_index(index_set, inplace=True)
    df[col][index_r] = value
    df.reset_index(level=0, inplace=True)
    updatedftoexcel1(df,sheet_name)


def updatedftoexcel1(df,sheetName):
    append_df_to_excel(file_path + '/Grist Optimizer changed.xlsx', df, sheet_name=sheetName,startrow=0,index=False)          
        


############################## Filling Days on Inventory 
##############################


def doiCal(val,list_):
#     print('val',val)
#     print('list_',len(list_))
    list_ = [0 if math.isnan(x) else x for x in list_]
#     list_ = list_.to_list()
    doi = 0
    if val > 0:
        for i in list_ :
            val = val - i
            if val < 0:
                break
            else :
                doi = doi + 1
    # print('DOI--',doi)
    return doi
    
def doiUpdate(ed):
    for i in range(len(ed)):
        for j in types:
            for k in plants:
                if ed['OH'+j+k][i] != '' :
                    doi = doiCal(ed['OH'+j+k][i],ed['R'+j+k][i+1:])
                    ed['DOI'+j+k][i] = doi

    return ed  



############################## Importing Excel 
##############################


ex = pd.ExcelFile(file_path + '/Inventory optimization working.xlsx')
ed = ex.parse('Sheet2')


plants = ['P1','P2','P3','P4','P5','P6','P7','P8','P9','P10']
types = ['C','L','S']
table_type = ['R','OH','DOI']

RCP = ed[['RCP1', 'RCP2', 'RCP3', 'RCP4', 'RCP5', 'RCP6', 'RCP7', 'RCP8', 'RCP9','RCP10']]
RLP = ed[['RLP1', 'RLP2', 'RLP3', 'RLP4', 'RLP5', 'RLP6', 'RLP7', 'RLP8', 'RLP9','RLP10']]
RSP = ed[['RSP1', 'RSP2', 'RSP3', 'RSP4', 'RSP5', 'RSP6', 'RSP7', 'RSP8', 'RSP9','RSP10']]
ed.columns[2:]



DOICP = ed[['id','OHCP1', 'OHCP2', 'OHCP3', 'OHCP4', 'OHCP5','OHCP6', 'OHCP7', 'OHCP8', 'OHCP9', 'OHCP10']]
DOILP = ed[['id','OHLP1', 'OHLP2', 'OHLP3','OHLP4', 'OHLP5', 'OHLP6', 'OHLP7', 'OHLP8', 'OHLP9', 'OHLP10']]
DOISP = ed[['id','OHSP1', 'OHSP2', 'OHSP3', 'OHSP4', 'OHSP5', 'OHSP6', 'OHSP7', 'OHSP8', 'OHSP9','OHSP10']]

## Calculating Aggregation
ed['RCAgg'] = ed['RCP1'].fillna(0) +ed['RCP2'].fillna(0)+ed['RCP3'].fillna(0)+ed['RCP4'].fillna(0)+ed['RCP5'].fillna(0)+ed['RCP6'].fillna(0)+ed['RCP7'].fillna(0)+ed['RCP8'].fillna(0)+ed['RCP9'].fillna(0)+ed['RCP10'].fillna(0)
ed['RLAgg'] = ed['RLP1'].fillna(0) +ed['RLP2'].fillna(0)+ed['RLP3'].fillna(0)+ed['RLP4'].fillna(0)+ed['RLP5'].fillna(0)+ed['RLP6'].fillna(0)+ed['RLP7'].fillna(0)+ed['RLP8'].fillna(0)+ed['RLP9'].fillna(0)+ed['RLP10'].fillna(0)
ed['RSAgg'] = ed['RSP1'].fillna(0) +ed['RSP2'].fillna(0)+ed['RSP3'].fillna(0)+ed['RSP4'].fillna(0)+ed['RSP5'].fillna(0)+ed['RSP6'].fillna(0)+ed['RSP7'].fillna(0)+ed['RSP8'].fillna(0)+ed['RSP9'].fillna(0)+ed['RSP10'].fillna(0)

##update DOI values
ed = doiUpdate(ed)

for i in range(len(ed)):
    for j in types:
        for k in plants:
            if ed['R'+j+k].fillna(0).sum() == 0 :
            	ed['DOI'+j+k] = 999999


##updating output to excel
updatedftoexcel(ed,'Sheet1')






# ############################# DoiGrstlvl
wheattype=['Canadian','SRW - US','LPRussian','LPGerman','LPFrench','LPArgentinaian']
planttype=["TCA","TCB","BL","PH","Warri","Calabar","Apapa","Ikorodu","Ilorin","Kano"]


RCAct=["RCActP1","RCActP2","RCActP3","RCActP4","RCActP5","RCActP6","RCActP7","RCActP8","RCActP9","RCActP10"]
RLAct=["RLActP1","RLActP2","RLActP3","RLActP4","RLActP5","RLActP6","RLActP7","RLActP8","RLActP9","RLActP10"]
RSAct=["RSActP1","RSActP2","RSActP3","RSActP4","RSActP5","RSActP6","RSActP7","RSActP8","RSActP9","RSActP10"]
RActualF=["RCActFlag","RLActFlag","RSActFlag"]
RLPType=["RLTypeP1","RLTypeP2","RLTypeP3","RLTypeP4","RLTypeP5","RLTypeP6","RLTypeP7","RLTypeP8","RLTypeP9","RLTypeP10"]


Lptype=["Rus","Ger","Fre","Arg"]



DOIC = ['DOICP1', 'DOICP2', 'DOICP3', 'DOICP4', 'DOICP5','DOICP6', 'DOICP7', 'DOICP8', 'DOICP9', 'DOICP10']
DOIL = ['DOILP1', 'DOILP2', 'DOILP3','DOILP4', 'DOILP5', 'DOILP6', 'DOILP7', 'DOILP8', 'DOILP9', 'DOILP10']
DOIS = ['DOISP1', 'DOISP2', 'DOISP3', 'DOISP4', 'DOISP5', 'DOISP6', 'DOISP7', 'DOISP8', 'DOISP9','DOISP10']

def DoiGrstlvl(edgrist):
    Cdoilist=[]
    Sdoilist=[]
    Ldoilist=[]

    
    for i in range(len(ed)):
        currentDT = datetime.datetime.now()
        d1=(currentDT.strftime("%Y-%m-%d 00:00:00"))
        # print('ed_date ',ed["Day"][i+1],type(ed["Day"][i+1]))
        # print('d1',d1,type(d1))
        if str(ed["Day"][i])==d1:
            for j in types:
                if j=="C":
                    for d in DOIC:
                        value=ed[d][i]
                        Cdoilist.append(value)

                if j=="S":
                    for d in DOIS:
                        value=ed[d][i]
                        Sdoilist.append(value)
                
                
                if j=="L":
                    for d in DOIL:
                        value=ed[d][i]
                        Ldoilist.append(value)
                    
    if len(Cdoilist) == 10:
        edgrist["Canadian"]= Cdoilist 
    if len(Sdoilist) == 10:
        edgrist["SRW - US"]= Sdoilist
    if len(Ldoilist) == 10:
        edgrist["LP"]= Ldoilist
    
    
    

    # print(edgrist)
    print("done")
    return edgrist
           
        
# DoiGrstlvl(edgrist)

exgrist = pd.ExcelFile(file_path + '/Grist Optimizer changed.xlsx')
edgrist = exgrist.parse('DOIGlvl')

# Actual filling after user inputs value
edgrist = DoiGrstlvl(edgrist)

##updating output to excel
updatedftoexcel1(edgrist,'DOIGlvl')
