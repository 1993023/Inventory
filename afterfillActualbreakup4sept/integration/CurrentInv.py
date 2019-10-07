import pandas as pd
import numpy as np
import math
import os
import datetime


file_path = os.path.dirname(os.path.abspath( __file__ ))


############################# Writing in excel 
#############################

def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False, 
                       **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    Parameters:
      filename : File path or existing ExcelWriter
                 (Example: '/path/to/file.xlsx')
      df : dataframe to save to workbook
      sheet_name : Name of sheet which will contain DataFrame.
                   (default: 'Sheet1')
      startrow : upper left cell row to dump data frame.
                 Per default (startrow=None) calculate the last row
                 in the existing DF and write to the next row...
      truncate_sheet : truncate (remove and recreate) [sheet_name]
                       before writing DataFrame to Excel file
      to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                        [can be dictionary]

    Returns: None
    """
    from openpyxl import load_workbook

    import pandas as pd

    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl')

    # Python 2.x: define [FileNotFoundError] exception if it doesn't exist 
    try:
        FileNotFoundError
    except NameError:
        FileNotFoundError = IOError


    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)

        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)

        # copy existing sheets
        writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()
    
    
def updateValExcel(sheet_name,df,index_set,index_r,col,value):
    df.set_index(index_set, inplace=True)
    df[col][index_r] = value
    df.reset_index(level=0, inplace=True)
    updatedftoexcel(df,sheet_name)
    
def updatedftoexcel(df,sheetName):
    append_df_to_excel(file_path + '/Inventory optimization working.xlsx', df, sheet_name=sheetName,startrow=0,index=False) 




############################## Importing Excel 
##############################


ex = pd.ExcelFile(file_path + '/Inventory optimization working.xlsx')
ed = ex.parse('Sheet2')

excmn = pd.ExcelFile(file_path + '/commonsheet.xlsx')
edcmn = excmn.parse('Sheet1')


plants = ['P1','P2','P3','P4','P5','P6','P7','P8','P9','P10']
types = ['C','L','S']
table_type = ['R','OH','DOI']

RCP = ed[['RCP1', 'RCP2', 'RCP3', 'RCP4', 'RCP5', 'RCP6', 'RCP7', 'RCP8', 'RCP9','RCP10']]
RLP = ed[['RLP1', 'RLP2', 'RLP3', 'RLP4', 'RLP5', 'RLP6', 'RLP7', 'RLP8', 'RLP9','RLP10']]
RSP = ed[['RSP1', 'RSP2', 'RSP3', 'RSP4', 'RSP5', 'RSP6', 'RSP7', 'RSP8', 'RSP9','RSP10']]
ed.columns[2:]



DOICP = ed[['id','OHCP1', 'OHCP2', 'OHCP3', 'OHCP4', 'OHCP5','OHCP6', 'OHCP7', 'OHCP8', 'OHCP9', 'OHCP10']]
DOILP = ed[['id','OHLP1', 'OHLP2', 'OHLP3','OHLP4', 'OHLP5', 'OHLP6', 'OHLP7', 'OHLP8', 'OHLP9', 'OHLP10']]
DOISP = ed[['id','OHSP1', 'OHSP2', 'OHSP3', 'OHSP4', 'OHSP5', 'OHSP6', 'OHSP7', 'OHSP8', 'OHSP9','OHSP10']]




# ############################# System_inv
wheattype=['Canadian','SRW - US','LPRussian','LPGerman','LPFrench','LPArgentinaian']
planttype=["TCA","TCB","BL","PH","Warri","Calabar","Apapa","Ikorodu","Ilorin","Kano"]


# CSys=["CSP1","CSP2","CSP3","CSP4","CSP5","CSP6","CSP7","CSP8","CSP9","CSP10"]
# LPSys=["LPSP1","LPSP2","LPSP3","LPSP4","LPSP5","LPSP6","LPSP7","LPSP8","LPSP9","LPSP10"]
# SSys=["SSP1","SSP2","SSP3","SSP4","SSP5","SSP6","SSP7","SSP8","SSP9","SSP10"]

def Sysinv(ed):
    Csyslist=[]
    LPsyslist=[]
    Ssyslist=[]

    for i in range(len(ed)):
        # currentDT = datetime.datetime.now()
        # d1=(currentDT.strftime("%Y-%m-%d 00:00:00"))
        # print('ed_date ',ed["Day"][i+1],type(ed["Day"][i+1]))
        # print('d1',d1,type(d1))
        
        d1=edcmn["DateSelc"][0]
        d1=str(d1)
        # print(d1)
        # print(type(d1))
        if str(ed["Day"][i])==d1:
            for j in types:
                if j=="C":
                    for m in range(len(edcmn)):
                        k=edcmn["INVCWRS"][m]
                        Csyslist.append(k)
                    for x in range(len(plants)):
                        ed["C"+"S"+plants[x]][i-1]=Csyslist[x]

                if j=="L":
                    for m in range(len(edcmn)):
                        k=edcmn["LpAgg"][m]
                        LPsyslist.append(k)
                    for x in range(len(plants)):
                        ed["LP"+"S"+plants[x]][i-1]=LPsyslist[x]
                
                if j=="S":
                    for m in range(len(edcmn)):
                        k=edcmn["INVSOFT WHEAT"][m]
                        Ssyslist.append(k)
                    for x in range(len(plants)):
                        ed["S"+"S"+plants[x]][i-1]=Ssyslist[x]
        
                          
                                             
       
    print("done")
    return ed
           
        
# Sysinv(ed)

# Actual filling after user inputs value
ed = Sysinv(ed)

##updating output to excel
updatedftoexcel(ed,'Sheet1')

