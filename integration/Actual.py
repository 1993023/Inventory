import pandas as pd
import numpy as np
import math
import os
import datetime

file_path = os.path.dirname(os.path.abspath( __file__ ))


############################## Writing in excel 
##############################

def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False, 
                       **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    Parameters:
      filename : File path or existing ExcelWriter
                 (Example: '/path/to/file.xlsx')
      df : dataframe to save to workbook
      sheet_name : Name of sheet which will contain DataFrame.
                   (default: 'Sheet1')
      startrow : upper left cell row to dump data frame.
                 Per default (startrow=None) calculate the last row
                 in the existing DF and write to the next row...
      truncate_sheet : truncate (remove and recreate) [sheet_name]
                       before writing DataFrame to Excel file
      to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                        [can be dictionary]

    Returns: None
    """
    from openpyxl import load_workbook

    import pandas as pd

    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl')

    # Python 2.x: define [FileNotFoundError] exception if it doesn't exist 
    try:
        FileNotFoundError
    except NameError:
        FileNotFoundError = IOError


    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)

        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)

        # copy existing sheets
        writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()
    
    
def updateValExcel(sheet_name,df,index_set,index_r,col,value):
    df.set_index(index_set, inplace=True)
    df[col][index_r] = value
    df.reset_index(level=0, inplace=True)
    updatedftoexcel(df,sheet_name)
    
def updatedftoexcel(df,sheetName):
    append_df_to_excel(file_path + '/Inventory optimization working.xlsx', df, sheet_name=sheetName,startrow=0,index=False) 


def updateValExcel1(sheet_name,df,index_set,index_r,col,value):
    df.set_index(index_set, inplace=True)
    df[col][index_r] = value
    df.reset_index(level=0, inplace=True)
    updatedftoexcel1(df,sheet_name)


def updatedftoexcel1(df,sheetName):
    append_df_to_excel(file_path + '/Grist Optimizer changed.xlsx', df, sheet_name=sheetName,startrow=0,index=False)          


############################## Actual Filling
##############################

def contriPlant(df,type_,start,end):
    list_a = []
    for k in afterskip:
        list_a.append(int(df['R'+type_+k][start:end].fillna(0).sum()))
#         print('Plant:',k,'--',df['R'+type_+k][start:end].fillna(0).sum())
    if list_a:
        return list_a
    else :
        return None
# list_a = contriPlant(ed,'C',3,8)
# print(list_a)


#########  def Actual Flag Value


def checkNegative1(list_ag,loc,def_inventory):
    list_ag = [0 if math.isnan(x) else x for x in list_ag]
#     print(len(list_ag),len(list_ag)-len(list_ag[loc:]))
    list_ag = list_ag[loc:]
    return_loc = 0
    for i in range(len(list_ag)):
        if def_inventory - list_ag[i] >= 0 :
            def_inventory = def_inventory -list_ag[i]
            return_loc = i
            # print(i,list_ag[i])
        else :
#             print(i,def_inventory -list_ag[i])
            break 
    if return_loc > 0:
#         list_ = contriPlant(df,j,loc,return_loc + loc+1)
        return return_loc + loc


def calculateRCAgg():
    rcagg=0
    for k in afterskip:
        rcagg=rcagg+ed['RC'+k].fillna(0)
    return rcagg


def calculateRLAgg():
    rlagg=0
    for k in afterskip:
        rlagg=rlagg+ed['RL'+k].fillna(0)
    return rlagg


def calculateRSAgg():
    rsagg=0
    for k in afterskip:
        rsagg=rsagg+ed['RS'+k].fillna(0)
    return rsagg

def actualReceived(ed):
    def_inventory_g = order_quantity
    for j in types:
        temp = None
#         for i in range(len(ed)):
        for i in range(len(ed)-1,-1,-1):
            for k in afterskip:
                if ed['R'+j+'ActFlag'][i] > 0 :
                    if j == 'C':
                        list_ag = ed['R'+j+'Agg']
                        # print("*************listagg ********************",list_ag)
                        print('Value',ed['R'+j+'ActFlag'][i],'Typej=',j,' Plantk=',k,' Rowi=',i)                 

    #                     ed['R'+j+'ActFlag'][i] = def_inventory_g 

                        t = checkNegative1(list_ag,i,ed['R'+j+'ActFlag'][i])
                        list_ = contriPlant(ed,j,i,t+1)
                        print("before distribution****************************************************",list_)
                        # print('list_---- Once sum******************************************* ',sum(list_))
                        bl=order_quantity-sum(list_)
                        ind=list_.index(max(list_))
                        list_[ind]=list_[ind]+bl
                        print("After  Adding balance distribution****************************************************",list_)

                        totaldistrib=[]
                        firstPortEndpos=[]
                        if port_name[0]=="Warri":
                            firstport=Warri_Port
                            pos=TWTvalue["Warri"]+1
                        if port_name[0]=="PH":
                            firstport=PH_Port
                            pos=TWTvalue["PH"]+1
                        if port_name[0]=="Lagos":
                            firstport=Lagosport
                            pos=TWTvalue["Lagos"]+1
                        if port_name[0]=="Port4":
                            firstport=PortFour
                            pos=TWTvalue["Port4"]+1
                        if port_name[0]=="Port5":
                            firstport=PortFive
                            pos=TWTvalue["Port5"]+1
                        
                        balance=0
                        for x in range(len(afterskip)):
                            # discharge_days = discharge[x]
                            # ed['R'+j+'Act'+afterskip[x]][i+discharge_days] = list_[x]
                            # pos=i
                            if afterskip[x] in firstport and list_[x]>0:
                                da=list_[x]
                                print("daaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaa",da)
                                da=da-balance
                                print("da ------------ balanceeeeeeeeeeeeeeeeeeeeeeeeeeeee",da)
                                while da>4500:
                                    print("pos**************************************",pos)
                                    # print("before daaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaa",da)
                                    ed['R'+j+'Act'+afterskip[x]][i+pos]=4500
                                    totaldistrib.append(4500)
                                    da=da-4500
                                    # print("after daaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaa",da)
                                    pos=pos+1
                                    firstPortEndpos.append(pos)
                                if da<4500 and da>=0:
                                    # print("less*********************************",da)
                                    ed['R'+j+'Act'+afterskip[x]][i+pos]=da
                                    totaldistrib.append(da)
                                    balance=4500-da
                                    # print("balance******************",balance)
                                    try:
                                        if list_[x+1]!=0:
                                            if afterskip[x+1] in firstport:
                                                if list_[x+1]>balance:
                                                    ed['R'+j+'Act'+afterskip[x+1]][i+pos]=balance
                                                    totaldistrib.append(balance)
                                                else:
                                                    ed['R'+j+'Act'+afterskip[x+1]][i+pos]=list_[x+1]
                                                    totaldistrib.append(list_[x+1])
                                                    balance=balance-list_[x+1]
                                                    
                                            else:
                                                ed['R'+j+'Act'+afterskip[x]][i+pos]=da
                                                # totaldistrib.append(da)

                                            pos=pos+1
                                            firstPortEndpos.append(pos)
                                            print("posssssssssssssssssss",pos)
                                        else:
                                            pos=pos+1
                                            firstPortEndpos.append(pos)
                                    except:
                                        pass
                            # else:
                            #     firstPortEndpos.append(1)
                        

                        secondPortEndpos=[]
                        if port_name[1]=="Warri":
                            secondPort=Warri_Port
                            pos=TWTvalue["Warri"]+firstPortEndpos[-1]
                        if port_name[1]=="PH":
                            secondPort=PH_Port
                            pos=TWTvalue["PH"]+firstPortEndpos[-1]
                        if port_name[1]=="Lagos":
                            secondPort=Lagosport
                            pos=TWTvalue["Lagos"]+firstPortEndpos[-1]
                        if port_name[1]=="Port4":
                            secondPort=PortFour
                            pos=TWTvalue["Port4"]+firstPortEndpos[-1]
                        if port_name[1]=="Port5":
                            secondPort=PortFive
                            pos=TWTvalue["Port5"]+firstPortEndpos[-1]
                        # pos=lagosEndpos[-1]+TWTvalue["PH"]
                        balance=0
                        for x in range(len(afterskip)):
                            if afterskip[x] in secondPort and list_[x]>0:
                                da=list_[x]
                                da=da-balance
                                while da>4500:
                                    print("pos**************************************",pos)
                                    ed['R'+j+'Act'+afterskip[x]][i+pos]=4500
                                    totaldistrib.append(4500)
                                    da=da-4500
                                    pos=pos+1
                                    secondPortEndpos.append(pos)
                                if da<4500 and da>=0:
                                    ed['R'+j+'Act'+afterskip[x]][i+pos]=da
                                    totaldistrib.append(da)
                                    balance=4500-da
                                    try:
                                        if list_[x+1]!=0:
                                            if afterskip[x+1] in secondPort:
                                                if list_[x+1]>balance:
                                                    ed['R'+j+'Act'+afterskip[x+1]][i+pos]=balance
                                                    totaldistrib.append(balance)
                                                else:
                                                    ed['R'+j+'Act'+afterskip[x+1]][i+pos]=list_[x+1]
                                                    totaldistrib.append(list_[x+1])
                                                    balance=balance-list_[x+1]
                                            else:
                                                ed['R'+j+'Act'+afterskip[x]][i+pos]=da
                                                # totaldistrib.append(da)
                                                
                                            pos=pos+1
                                            secondPortEndpos.append(pos)
                                        else:
                                            pos=pos+1
                                            secondPortEndpos.append(pos)
                                    except:
                                        pass
                        
                            else:
                                secondPortEndpos=firstPortEndpos
                        
                        thirdPortEndpos=[]
                        if port_name[2]=="Warri":
                            thirdPort=Warri_Port
                            pos=TWTvalue["Warri"]+secondPortEndpos[-1]
                        if port_name[2]=="PH":
                            thirdPort=PH_Port
                            pos=TWTvalue["PH"]+secondPortEndpos[-1]
                        if port_name[2]=="Lagos":
                            thirdPort=Lagosport
                            pos=TWTvalue["Lagos"]+secondPortEndpos[-1]
                        if port_name[2]=="Port4":
                            thirdPort=PortFour
                            pos=TWTvalue["Port4"]+secondPortEndpos[-1]
                        if port_name[2]=="Port5":
                            thirdPort=PortFive
                            pos=TWTvalue["Port5"]+secondPortEndpos[-1]
                        # pos=lagosEndpos[-1]+TWTvalue["PH"]
                        balance=0
                        for x in range(len(afterskip)):
                            if afterskip[x] in thirdPort and list_[x]>0:
                                da=list_[x]
                                da=da-balance
                                while da>4500:
                                    ed['R'+j+'Act'+afterskip[x]][i+pos]=4500
                                    totaldistrib.append(4500)
                                    da=da-4500
                                    pos=pos+1
                                    thirdPortEndpos.append(pos)
                                if da<4500 and da>=0:
                                    ed['R'+j+'Act'+afterskip[x]][i+pos]=da
                                    totaldistrib.append(da)
                                    balance=4500-da
                                    try:
                                        if list_[x+1]!=0:
                                            if afterskip[x+1] in thirdPort:
                                                if list_[x+1]>balance:
                                                    ed['R'+j+'Act'+afterskip[x+1]][i+pos]=balance
                                                    totaldistrib.append(balance)
                                                else:
                                                    ed['R'+j+'Act'+afterskip[x+1]][i+pos]=list_[x+1]
                                                    totaldistrib.append(list_[x+1])
                                                    balance=balance-list_[x+1]
                                            else:
                                                ed['R'+j+'Act'+afterskip[x]][i+pos]=da
                                                # totaldistrib.append(da)
                                            pos=pos+1
                                            thirdPortEndpos.append(pos)
                                        else:
                                            pos=pos+1
                                            thirdPortEndpos.append(pos)
                                    except:
                                        pass
                            else:
                                thirdPortEndpos=secondPortEndpos

                        
                        fourthPortEndpos=[]
                        if port_name[3]=="Warri":
                            fourthPort=Warri_Port
                            pos=TWTvalue["Warri"]+thirdPortEndpos[-1]
                        if port_name[3]=="PH":
                            fourthPort=PH_Port
                            pos=TWTvalue["PH"]+thirdPortEndpos[-1]
                        if port_name[3]=="Lagos":
                            fourthPort=Lagosport
                            pos=TWTvalue["Lagos"]+thirdPortEndpos[-1]
                        if port_name[3]=="Port4":
                            fourthPort=PortFour
                            pos=TWTvalue["Port4"]+thirdPortEndpos[-1]
                        if port_name[3]=="Port5":
                            fourthPort=PortFive
                            pos=TWTvalue["Port5"]+thirdPortEndpos[-1]
                        # pos=lagosEndpos[-1]+TWTvalue["PH"]
                        balance=0
                        for x in range(len(afterskip)):
                            if afterskip[x] in fourthPort and list_[x]>0:
                                da=list_[x]
                                da=da-balance
                                while da>4500:
                                    ed['R'+j+'Act'+afterskip[x]][i+pos]=4500
                                    totaldistrib.append(4500)
                                    da=da-4500
                                    pos=pos+1
                                    fourthPortEndpos.append(pos)
                                if da<4500 and da>=0:
                                    ed['R'+j+'Act'+afterskip[x]][i+pos]=da
                                    totaldistrib.append(da)
                                    balance=4500-da
                                    try:
                                        if list_[x+1]!=0:
                                            if afterskip[x+1] in fourthPort:
                                                if list_[x+1]>balance:
                                                    ed['R'+j+'Act'+afterskip[x+1]][i+pos]=balance
                                                    totaldistrib.append(balance)
                                                else:
                                                    ed['R'+j+'Act'+afterskip[x+1]][i+pos]=list_[x+1]
                                                    totaldistrib.append(list_[x+1])
                                                    balance=balance-list_[x+1]
                                            else:
                                                ed['R'+j+'Act'+afterskip[x]][i+pos]=da
                                                # totaldistrib.append(da)
                                            pos=pos+1
                                            fourthPortEndpos.append(pos)
                                        else:
                                            pos=pos+1
                                            fourthPortEndpos.append(pos)
                                    except:
                                        pass
                            else:
                                fourthPortEndpos=thirdPortEndpos

                        
                        
                        fifthPortEndpos=[]
                        if port_name[4]=="Warri":
                            fifthPort=Warri_Port
                            pos=TWTvalue["Warri"]+fourthPortEndpos[-1]
                        if port_name[4]=="PH":
                            fifthPort=PH_Port
                            pos=TWTvalue["PH"]+fourthPortEndpos[-1]
                        if port_name[4]=="Lagos":
                            fifthPort=Lagosport
                            pos=TWTvalue["Lagos"]+fourthPortEndpos[-1]
                        if port_name[4]=="Port4":
                            fifthPort=PortFour
                            pos=TWTvalue["Port4"]+fourthPortEndpos[-1]
                        if port_name[4]=="Port5":
                            fifthPort=PortFive
                            pos=TWTvalue["Port5"]+fourthPortEndpos[-1]
                        # pos=lagosEndpos[-1]+TWTvalue["PH"]
                        balance=0
                        count=0
                        for x in range(len(afterskip)):
                            if afterskip[x] in fifthPort and list_[x]>0:
                                # if count==0:
                                #     da=totalbalance
                                # else:
                                #     da=list_[x]
                                # count=count+1
                                da=list_[x]
                                da=da-balance
                                while da>4500:
                                    ed['R'+j+'Act'+afterskip[x]][i+pos]=4500
                                    totaldistrib.append(4500)
                                    da=da-4500
                                    pos=pos+1
                                    fifthPortEndpos.append(pos)
                                if da<4500 and da>=0:
                                    ed['R'+j+'Act'+afterskip[x]][i+pos]=da
                                    totaldistrib.append(da)
                                    balance=4500-da
                                    try:
                                        if list_[x+1]!=0:
                                            if afterskip[x+1] in fifthPort:
                                                if list_[x+1]>balance:
                                                    ed['R'+j+'Act'+afterskip[x+1]][i+pos]=balance
                                                    totaldistrib.append(balance)
                                                else:
                                                    ed['R'+j+'Act'+afterskip[x+1]][i+pos]=list_[x+1]
                                                    totaldistrib.append(list_[x+1])
                                                    balance=balance-list_[x+1]
                                            else:
                                                ed['R'+j+'Act'+afterskip[x]][i+pos]=da
                                                # totaldistrib.append(da)
                                            pos=pos+1
                                            fifthPortEndpos.append(pos)
                                        else:
                                            pos=pos+1
                                            fifthPortEndpos.append(pos)
                                    except:
                                        pass
                            else:
                                fifthPortEndpos=fourthPortEndpos
                                   


                        print("final total distibuted is here********************************",(totaldistrib))
                        print("sum is here********************************",sum(totaldistrib))
                        totalbalance=order_quantity-sum(totaldistrib)
                        print("final total balance is here********************************",(totalbalance))
                        print('list_---- Once ',list_)
                        print('list_---- Once ',sum(list_))



    #                     ed['R'+j+'Act'+k][i] = ed['R'+j+k][i:t].fillna(0).sum()

                        ed['R'+j+'ActFlag'][t+1] =  def_inventory_g
                        m = t + 1
                        while m <= len(ed):
                            print('m',m)
                            #[m+1:]

                            a = checkNegative1(list_ag,m,def_inventory_g)

    #                         a ,list_ = checkNegative1(ed,j,list_ag,m,def_inventory_g)

                            print('a-----',a)
                            if a == None :
                                m = len(ed) + 1
                            else : 
#                                 list_ = contriPlant(ed,j,m,a+1)
#                                 for x in range(len(plants)):
#                                     ed['R'+j+'Act'+plants[x]][m] = list_[x] 
                                ed['R'+j+'ActFlag'][a+1] = def_inventory_g
                                m = a + 1

                        temp = 'Found'
                        #print('break3-----------')
#                         print('**********','R'+j+'Flag',len(ed['R'+j+'Flag'][i+1:len(ed)]),'R'+j+'ActFlag',len(ed['R'+j+'ActFlag'][i+1:len(ed)]))
                        ed['R'+j+'Flag'][i+1:len(ed)] = ed['R'+j+'ActFlag'][i+1:len(ed)]
                        temp1 = ed['R'+j+'ActFlag'][0:i+1] 
                        ed['R'+j+'ActFlag'] = np.NAN
                        ed['R'+j+'ActFlag'] = temp1
                        break

                    if j == 'L':
                        list_ag = ed['R'+j+'Agg']
                        print('Value',ed['R'+j+'ActFlag'][i],'Typej=',j,' Plantk=',k,' Rowi=',i)                 

    #                     ed['R'+j+'ActFlag'][i] = def_inventory_g 

                        t = checkNegative1(list_ag,i,ed['R'+j+'ActFlag'][i])
                        list_ = contriPlant(ed,j,i,t+1)
                        LPType=ed["RLType"][i]
                        # for x in range(len(afterskip)):
                        #     discharge_days = discharge[x]
                        #     ed['R'+j+'Act'+afterskip[x]][i+discharge_days] = list_[x] 
                        #     ed['R'+j+'Type'+afterskip[x]][i+discharge_days]=LPType
                        # bl=order_quantity-sum(list_)
                        # ind=list_.index(max(list_))
                        # list_[ind]=list_[ind]+bl
                        # ******recent Adding****
                        print("before distribution****************************************************",list_)
                        # print('list_---- Once sum******************************************* ',sum(list_))
                        bl=order_quantity-sum(list_)
                        ind=list_.index(max(list_))
                        list_[ind]=list_[ind]+bl
                        print("After  Adding balance distribution****************************************************",list_)

                        totaldistrib=[]
                        firstPortEndpos=[]
                        if port_name[0]=="Warri":
                            firstport=Warri_Port
                            pos=TWTvalue["Warri"]+1
                        if port_name[0]=="PH":
                            firstport=PH_Port
                            pos=TWTvalue["PH"]+1
                        if port_name[0]=="Lagos":
                            firstport=Lagosport
                            pos=TWTvalue["Lagos"]+1
                        if port_name[0]=="Port4":
                            firstport=PortFour
                            pos=TWTvalue["Port4"]+1
                        if port_name[0]=="Port5":
                            firstport=PortFive
                            pos=TWTvalue["Port5"]+1
                        
                        balance=0
                        for x in range(len(afterskip)):
                            # discharge_days = discharge[x]
                            # ed['R'+j+'Act'+afterskip[x]][i+discharge_days] = list_[x]
                            # pos=i
                            if afterskip[x] in firstport and list_[x]>0:
                                da=list_[x]
                                print("daaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaa",da)
                                da=da-balance
                                print("da ------------ balanceeeeeeeeeeeeeeeeeeeeeeeeeeeee",da)
                                while da>4500:
                                    print("pos**************************************",pos)
                                    # print("before daaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaa",da)
                                    ed['R'+j+'Act'+afterskip[x]][i+pos]=4500
                                    ed['R'+j+'Type'+afterskip[x]][i+pos]=LPType
                                    totaldistrib.append(4500)
                                    da=da-4500
                                    # print("after daaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaa",da)
                                    pos=pos+1
                                    firstPortEndpos.append(pos)
                                if da<4500 and da>=0:
                                    # print("less*********************************",da)
                                    ed['R'+j+'Act'+afterskip[x]][i+pos]=da
                                    ed['R'+j+'Type'+afterskip[x]][i+pos]=LPType
                                    totaldistrib.append(da)
                                    balance=4500-da
                                    # print("balance******************",balance)
                                    try:
                                        if list_[x+1]!=0:
                                            if afterskip[x+1] in firstport:
                                                if list_[x+1]>balance:
                                                    ed['R'+j+'Act'+afterskip[x+1]][i+pos]=balance
                                                    ed['R'+j+'Type'+afterskip[x+1]][i+pos]=LPType
                                                    totaldistrib.append(balance)
                                                else:
                                                    ed['R'+j+'Act'+afterskip[x+1]][i+pos]=list_[x+1]
                                                    ed['R'+j+'Type'+afterskip[x+1]][i+pos]=LPType
                                                    totaldistrib.append(list_[x+1])
                                                    balance=balance-list_[x+1]
                                                    
                                            else:
                                                ed['R'+j+'Act'+afterskip[x]][i+pos]=da
                                                ed['R'+j+'Type'+afterskip[x]][i+pos]=LPType
                                                # totaldistrib.append(da)

                                            pos=pos+1
                                            firstPortEndpos.append(pos)
                                            print("posssssssssssssssssss",pos)
                                        else:
                                            pos=pos+1
                                            firstPortEndpos.append(pos)
                                    except:
                                        pass
                            # else:
                            #     firstPortEndpos.append(1)
                        

                        secondPortEndpos=[]
                        if port_name[1]=="Warri":
                            secondPort=Warri_Port
                            pos=TWTvalue["Warri"]+firstPortEndpos[-1]
                        if port_name[1]=="PH":
                            secondPort=PH_Port
                            pos=TWTvalue["PH"]+firstPortEndpos[-1]
                        if port_name[1]=="Lagos":
                            secondPort=Lagosport
                            pos=TWTvalue["Lagos"]+firstPortEndpos[-1]
                        if port_name[1]=="Port4":
                            secondPort=PortFour
                            pos=TWTvalue["Port4"]+firstPortEndpos[-1]
                        if port_name[1]=="Port5":
                            secondPort=PortFive
                            pos=TWTvalue["Port5"]+firstPortEndpos[-1]
                        # pos=lagosEndpos[-1]+TWTvalue["PH"]
                        balance=0
                        for x in range(len(afterskip)):
                            if afterskip[x] in secondPort and list_[x]>0:
                                da=list_[x]
                                da=da-balance
                                while da>4500:
                                    print("pos**************************************",pos)
                                    ed['R'+j+'Act'+afterskip[x]][i+pos]=4500
                                    ed['R'+j+'Type'+afterskip[x]][i+pos]=LPType
                                    totaldistrib.append(4500)
                                    da=da-4500
                                    pos=pos+1
                                    secondPortEndpos.append(pos)
                                if da<4500 and da>=0:
                                    ed['R'+j+'Act'+afterskip[x]][i+pos]=da
                                    ed['R'+j+'Type'+afterskip[x]][i+pos]=LPType
                                    totaldistrib.append(da)
                                    balance=4500-da
                                    try:
                                        if list_[x+1]!=0:
                                            if afterskip[x+1] in secondPort:
                                                if list_[x+1]>balance:
                                                    ed['R'+j+'Act'+afterskip[x+1]][i+pos]=balance
                                                    ed['R'+j+'Type'+afterskip[x+1]][i+pos]=LPType
                                                    totaldistrib.append(balance)
                                                else:
                                                    ed['R'+j+'Act'+afterskip[x+1]][i+pos]=list_[x+1]
                                                    ed['R'+j+'Type'+afterskip[x+1]][i+pos]=LPType
                                                    totaldistrib.append(list_[x+1])
                                                    balance=balance-list_[x+1]
                                            else:
                                                ed['R'+j+'Act'+afterskip[x]][i+pos]=da
                                                ed['R'+j+'Type'+afterskip[x]][i+pos]=LPType
                                                # totaldistrib.append(da)
                                                
                                            pos=pos+1
                                            secondPortEndpos.append(pos)
                                        else:
                                            pos=pos+1
                                            secondPortEndpos.append(pos)
                                    except:
                                        pass
                        
                            else:
                                secondPortEndpos=firstPortEndpos
                        
                        thirdPortEndpos=[]
                        if port_name[2]=="Warri":
                            thirdPort=Warri_Port
                            pos=TWTvalue["Warri"]+secondPortEndpos[-1]
                        if port_name[2]=="PH":
                            thirdPort=PH_Port
                            pos=TWTvalue["PH"]+secondPortEndpos[-1]
                        if port_name[2]=="Lagos":
                            thirdPort=Lagosport
                            pos=TWTvalue["Lagos"]+secondPortEndpos[-1]
                        if port_name[2]=="Port4":
                            thirdPort=PortFour
                            pos=TWTvalue["Port4"]+secondPortEndpos[-1]
                        if port_name[2]=="Port5":
                            thirdPort=PortFive
                            pos=TWTvalue["Port5"]+secondPortEndpos[-1]
                        # pos=lagosEndpos[-1]+TWTvalue["PH"]
                        balance=0
                        for x in range(len(afterskip)):
                            if afterskip[x] in thirdPort and list_[x]>0:
                                da=list_[x]
                                da=da-balance
                                while da>4500:
                                    ed['R'+j+'Act'+afterskip[x]][i+pos]=4500
                                    ed['R'+j+'Type'+afterskip[x]][i+pos]=LPType
                                    totaldistrib.append(4500)
                                    da=da-4500
                                    pos=pos+1
                                    thirdPortEndpos.append(pos)
                                if da<4500 and da>=0:
                                    ed['R'+j+'Act'+afterskip[x]][i+pos]=da
                                    ed['R'+j+'Type'+afterskip[x]][i+pos]=LPType
                                    totaldistrib.append(da)
                                    balance=4500-da
                                    try:
                                        if list_[x+1]!=0:
                                            if afterskip[x+1] in thirdPort:
                                                if list_[x+1]>balance:
                                                    ed['R'+j+'Act'+afterskip[x+1]][i+pos]=balance
                                                    ed['R'+j+'Type'+afterskip[x+1]][i+pos]=LPType
                                                    totaldistrib.append(balance)
                                                else:
                                                    ed['R'+j+'Act'+afterskip[x+1]][i+pos]=list_[x+1]
                                                    ed['R'+j+'Type'+afterskip[x+1]][i+pos]=LPType
                                                    totaldistrib.append(list_[x+1])
                                                    balance=balance-list_[x+1]
                                            else:
                                                ed['R'+j+'Act'+afterskip[x]][i+pos]=da
                                                ed['R'+j+'Type'+afterskip[x]][i+pos]=LPType
                                                # totaldistrib.append(da)
                                            pos=pos+1
                                            thirdPortEndpos.append(pos)
                                        else:
                                            pos=pos+1
                                            thirdPortEndpos.append(pos)
                                    except:
                                        pass
                            else:
                                thirdPortEndpos=secondPortEndpos

                        
                        fourthPortEndpos=[]
                        if port_name[3]=="Warri":
                            fourthPort=Warri_Port
                            pos=TWTvalue["Warri"]+thirdPortEndpos[-1]
                        if port_name[3]=="PH":
                            fourthPort=PH_Port
                            pos=TWTvalue["PH"]+thirdPortEndpos[-1]
                        if port_name[3]=="Lagos":
                            fourthPort=Lagosport
                            pos=TWTvalue["Lagos"]+thirdPortEndpos[-1]
                        if port_name[3]=="Port4":
                            fourthPort=PortFour
                            pos=TWTvalue["Port4"]+thirdPortEndpos[-1]
                        if port_name[3]=="Port5":
                            fourthPort=PortFive
                            pos=TWTvalue["Port5"]+thirdPortEndpos[-1]
                        # pos=lagosEndpos[-1]+TWTvalue["PH"]
                        balance=0
                        for x in range(len(afterskip)):
                            if afterskip[x] in fourthPort and list_[x]>0:
                                da=list_[x]
                                da=da-balance
                                while da>4500:
                                    ed['R'+j+'Act'+afterskip[x]][i+pos]=4500
                                    ed['R'+j+'Type'+afterskip[x]][i+pos]=LPType
                                    totaldistrib.append(4500)
                                    da=da-4500
                                    pos=pos+1
                                    fourthPortEndpos.append(pos)
                                if da<4500 and da>=0:
                                    ed['R'+j+'Act'+afterskip[x]][i+pos]=da
                                    ed['R'+j+'Type'+afterskip[x]][i+pos]=LPType
                                    totaldistrib.append(da)
                                    balance=4500-da
                                    try:
                                        if list_[x+1]!=0:
                                            if afterskip[x+1] in fourthPort:
                                                if list_[x+1]>balance:
                                                    ed['R'+j+'Act'+afterskip[x+1]][i+pos]=balance
                                                    ed['R'+j+'Type'+afterskip[x+1]][i+pos]=LPType
                                                    totaldistrib.append(balance)
                                                else:
                                                    ed['R'+j+'Act'+afterskip[x+1]][i+pos]=list_[x+1]
                                                    ed['R'+j+'Type'+afterskip[x+1]][i+pos]=LPType
                                                    totaldistrib.append(list_[x+1])
                                                    balance=balance-list_[x+1]
                                            else:
                                                ed['R'+j+'Act'+afterskip[x]][i+pos]=da
                                                ed['R'+j+'Type'+afterskip[x]][i+pos]=LPType
                                                # totaldistrib.append(da)
                                            pos=pos+1
                                            fourthPortEndpos.append(pos)
                                        else:
                                            pos=pos+1
                                            fourthPortEndpos.append(pos)
                                    except:
                                        pass
                            else:
                                fourthPortEndpos=thirdPortEndpos

                        
                        
                        fifthPortEndpos=[]
                        if port_name[4]=="Warri":
                            fifthPort=Warri_Port
                            pos=TWTvalue["Warri"]+fourthPortEndpos[-1]
                        if port_name[4]=="PH":
                            fifthPort=PH_Port
                            pos=TWTvalue["PH"]+fourthPortEndpos[-1]
                        if port_name[4]=="Lagos":
                            fifthPort=Lagosport
                            pos=TWTvalue["Lagos"]+fourthPortEndpos[-1]
                        if port_name[4]=="Port4":
                            fifthPort=PortFour
                            pos=TWTvalue["Port4"]+fourthPortEndpos[-1]
                        if port_name[4]=="Port5":
                            fifthPort=PortFive
                            pos=TWTvalue["Port5"]+fourthPortEndpos[-1]
                        # pos=lagosEndpos[-1]+TWTvalue["PH"]
                        balance=0
                        count=0
                        for x in range(len(afterskip)):
                            if afterskip[x] in fifthPort and list_[x]>0:
                                # if count==0:
                                #     da=totalbalance
                                # else:
                                #     da=list_[x]
                                # count=count+1
                                da=list_[x]
                                da=da-balance
                                while da>4500:
                                    ed['R'+j+'Act'+afterskip[x]][i+pos]=4500
                                    ed['R'+j+'Type'+afterskip[x]][i+pos]=LPType
                                    totaldistrib.append(4500)
                                    da=da-4500
                                    pos=pos+1
                                    fifthPortEndpos.append(pos)
                                if da<4500 and da>=0:
                                    ed['R'+j+'Act'+afterskip[x]][i+pos]=da
                                    ed['R'+j+'Type'+afterskip[x]][i+pos]=LPType
                                    totaldistrib.append(da)
                                    balance=4500-da
                                    try:
                                        if list_[x+1]!=0:
                                            if afterskip[x+1] in fifthPort:
                                                if list_[x+1]>balance:
                                                    ed['R'+j+'Act'+afterskip[x+1]][i+pos]=balance
                                                    ed['R'+j+'Type'+afterskip[x+1]][i+pos]=LPType
                                                    totaldistrib.append(balance)
                                                else:
                                                    ed['R'+j+'Act'+afterskip[x+1]][i+pos]=list_[x+1]
                                                    ed['R'+j+'Type'+afterskip[x+1]][i+pos]=LPType
                                                    totaldistrib.append(list_[x+1])
                                                    balance=balance-list_[x+1]
                                            else:
                                                ed['R'+j+'Act'+afterskip[x]][i+pos]=da
                                                ed['R'+j+'Type'+afterskip[x]][i+pos]=LPType
                                                # totaldistrib.append(da)
                                            pos=pos+1
                                            fifthPortEndpos.append(pos)
                                        else:
                                            pos=pos+1
                                            fifthPortEndpos.append(pos)
                                    except:
                                        pass
                            else:
                                fifthPortEndpos=fourthPortEndpos


                        print("final total distibuted is here********************************",(totaldistrib))
                        print("sum is here********************************",sum(totaldistrib))
                        totalbalance=order_quantity-sum(totaldistrib)
                        print("final total balance is here********************************",(totalbalance))
                        print('list_---- Once ',list_)


    #                     ed['R'+j+'Act'+k][i] = ed['R'+j+k][i:t].fillna(0).sum()

                        ed['R'+j+'ActFlag'][t+1] =  def_inventory_g
                        m = t + 1
                        while m <= len(ed):
                            print('m',m)
                            #[m+1:]

                            a = checkNegative1(list_ag,m,def_inventory_g)

    #                         a ,list_ = checkNegative1(ed,j,list_ag,m,def_inventory_g)

                            print('a-----',a)
                            if a == None :
                                m = len(ed) + 1
                            else : 
#                                 list_ = contriPlant(ed,j,m,a+1)
#                                 for x in range(len(plants)):
#                                     ed['R'+j+'Act'+plants[x]][m] = list_[x] 
                                ed['R'+j+'ActFlag'][a+1] = def_inventory_g
                                m = a + 1

                        temp = 'Found'
                        #print('break3-----------')
#                         print('**********','R'+j+'Flag',len(ed['R'+j+'Flag'][i+1:len(ed)]),'R'+j+'ActFlag',len(ed['R'+j+'ActFlag'][i+1:len(ed)]))
                        ed['R'+j+'Flag'][i+1:len(ed)] = ed['R'+j+'ActFlag'][i+1:len(ed)]
                        temp1 = ed['R'+j+'ActFlag'][0:i+1] 
                        ed['R'+j+'ActFlag'] = np.NAN
                        ed['R'+j+'ActFlag'] = temp1
                        break

                    if j == 'S':
                        list_ag = ed['R'+j+'Agg']
                        print('Value',ed['R'+j+'ActFlag'][i],'Typej=',j,' Plantk=',k,' Rowi=',i)                 

    #                     ed['R'+j+'ActFlag'][i] = def_inventory_g 

                        t = checkNegative1(list_ag,i,ed['R'+j+'ActFlag'][i])
                        list_ = contriPlant(ed,j,i,t+1)
                        # for x in range(len(afterskip)):
                        #     discharge_days = discharge[x]
                        #     ed['R'+j+'Act'+afterskip[x]][i+discharge_days] = list_[x] 
                        # bl=order_quantity-sum(list_)
                        # ind=list_.index(max(list_))
                        # list_[ind]=list_[ind]+bl
                        # ***********recent Adding****

                        print("before distribution****************************************************",list_)
                        # print('list_---- Once sum******************************************* ',sum(list_))
                        bl=order_quantity-sum(list_)
                        ind=list_.index(max(list_))
                        list_[ind]=list_[ind]+bl
                        print("After  Adding balance distribution****************************************************",list_)

                        totaldistrib=[]
                        firstPortEndpos=[]
                        if port_name[0]=="Warri":
                            firstport=Warri_Port
                            pos=TWTvalue["Warri"]+1
                        if port_name[0]=="PH":
                            firstport=PH_Port
                            pos=TWTvalue["PH"]+1
                        if port_name[0]=="Lagos":
                            firstport=Lagosport
                            pos=TWTvalue["Lagos"]+1
                        if port_name[0]=="Port4":
                            firstport=PortFour
                            pos=TWTvalue["Port4"]+1
                        if port_name[0]=="Port5":
                            firstport=PortFive
                            pos=TWTvalue["Port5"]+1
                        
                        balance=0
                        for x in range(len(afterskip)):
                            # discharge_days = discharge[x]
                            # ed['R'+j+'Act'+afterskip[x]][i+discharge_days] = list_[x]
                            # pos=i
                            if afterskip[x] in firstport and list_[x]>0:
                                da=list_[x]
                                print("daaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaa",da)
                                da=da-balance
                                print("da ------------ balanceeeeeeeeeeeeeeeeeeeeeeeeeeeee",da)
                                while da>4500:
                                    print("pos**************************************",pos)
                                    # print("before daaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaa",da)
                                    if ed['R'+j+'Act'+afterskip[x]][i+pos]>0:
                                        ed['R'+j+'Act'+afterskip[x]][i+pos]=ed['R'+j+'Act'+afterskip[x]][i+pos]+4500
                                    else:
                                        ed['R'+j+'Act'+afterskip[x]][i+pos]=4500
                                    totaldistrib.append(4500)
                                    da=da-4500
                                    # print("after daaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaa",da)
                                    pos=pos+1
                                    firstPortEndpos.append(pos)
                                if da<4500 and da>=0:
                                    # print("less*********************************",da)
                                    ed['R'+j+'Act'+afterskip[x]][i+pos]=da
                                    totaldistrib.append(da)
                                    balance=4500-da
                                    # print("balance******************",balance)
                                    try:
                                        if list_[x+1]!=0:
                                            if afterskip[x+1] in firstport:
                                                if list_[x+1]>balance:
                                                    ed['R'+j+'Act'+afterskip[x+1]][i+pos]=balance
                                                    totaldistrib.append(balance)
                                                else:
                                                    ed['R'+j+'Act'+afterskip[x+1]][i+pos]=list_[x+1]
                                                    totaldistrib.append(list_[x+1])
                                                    balance=balance-list_[x+1]
                                                    
                                            else:
                                                ed['R'+j+'Act'+afterskip[x]][i+pos]=da
                                                # totaldistrib.append(da)

                                            pos=pos+1
                                            firstPortEndpos.append(pos)
                                            print("posssssssssssssssssss",pos)
                                        else:
                                            pos=pos+1
                                            firstPortEndpos.append(pos)
                                    except:
                                        pass
                            else:
                                firstPortEndpos.append(1)
                        

                        secondPortEndpos=[]
                        if port_name[1]=="Warri":
                            secondPort=Warri_Port
                            pos=TWTvalue["Warri"]+firstPortEndpos[-1]
                        if port_name[1]=="PH":
                            secondPort=PH_Port
                            pos=TWTvalue["PH"]+firstPortEndpos[-1]
                        if port_name[1]=="Lagos":
                            secondPort=Lagosport
                            pos=TWTvalue["Lagos"]+firstPortEndpos[-1]
                        if port_name[1]=="Port4":
                            secondPort=PortFour
                            pos=TWTvalue["Port4"]+firstPortEndpos[-1]
                        if port_name[1]=="Port5":
                            secondPort=PortFive
                            pos=TWTvalue["Port5"]+firstPortEndpos[-1]
                        # pos=lagosEndpos[-1]+TWTvalue["PH"]
                        balance=0
                        for x in range(len(afterskip)):
                            if afterskip[x] in secondPort and list_[x]>0:
                                da=list_[x]
                                da=da-balance
                                while da>4500:
                                    print("pos**************************************",pos)
                                    ed['R'+j+'Act'+afterskip[x]][i+pos]=4500
                                    totaldistrib.append(4500)
                                    da=da-4500
                                    pos=pos+1
                                    secondPortEndpos.append(pos)
                                if da<4500 and da>=0:
                                    ed['R'+j+'Act'+afterskip[x]][i+pos]=da
                                    totaldistrib.append(da)
                                    balance=4500-da
                                    try:
                                        if list_[x+1]!=0:
                                            if afterskip[x+1] in secondPort:
                                                if list_[x+1]>balance:
                                                    ed['R'+j+'Act'+afterskip[x+1]][i+pos]=balance
                                                    totaldistrib.append(balance)
                                                else:
                                                    ed['R'+j+'Act'+afterskip[x+1]][i+pos]=list_[x+1]
                                                    totaldistrib.append(list_[x+1])
                                                    balance=balance-list_[x+1]
                                            else:
                                                ed['R'+j+'Act'+afterskip[x]][i+pos]=da
                                                # totaldistrib.append(da)
                                                
                                            pos=pos+1
                                            secondPortEndpos.append(pos)
                                        else:
                                            pos=pos+1
                                            secondPortEndpos.append(pos)
                                    except:
                                        pass
                        
                            else:
                                secondPortEndpos=firstPortEndpos
                        
                        thirdPortEndpos=[]
                        if port_name[2]=="Warri":
                            thirdPort=Warri_Port
                            pos=TWTvalue["Warri"]+secondPortEndpos[-1]
                        if port_name[2]=="PH":
                            thirdPort=PH_Port
                            pos=TWTvalue["PH"]+secondPortEndpos[-1]
                        if port_name[2]=="Lagos":
                            thirdPort=Lagosport
                            pos=TWTvalue["Lagos"]+secondPortEndpos[-1]
                        if port_name[2]=="Port4":
                            thirdPort=PortFour
                            pos=TWTvalue["Port4"]+secondPortEndpos[-1]
                        if port_name[2]=="Port5":
                            thirdPort=PortFive
                            pos=TWTvalue["Port5"]+secondPortEndpos[-1]
                        # pos=lagosEndpos[-1]+TWTvalue["PH"]
                        balance=0
                        for x in range(len(afterskip)):
                            if afterskip[x] in thirdPort and list_[x]>0:
                                da=list_[x]
                                da=da-balance
                                while da>4500:
                                    ed['R'+j+'Act'+afterskip[x]][i+pos]=4500
                                    totaldistrib.append(4500)
                                    da=da-4500
                                    pos=pos+1
                                    thirdPortEndpos.append(pos)
                                if da<4500 and da>=0:
                                    ed['R'+j+'Act'+afterskip[x]][i+pos]=da
                                    totaldistrib.append(da)
                                    balance=4500-da
                                    try:
                                        if list_[x+1]!=0:
                                            if afterskip[x+1] in thirdPort:
                                                if list_[x+1]>balance:
                                                    ed['R'+j+'Act'+afterskip[x+1]][i+pos]=balance
                                                    totaldistrib.append(balance)
                                                else:
                                                    ed['R'+j+'Act'+afterskip[x+1]][i+pos]=list_[x+1]
                                                    totaldistrib.append(list_[x+1])
                                                    balance=balance-list_[x+1]
                                            else:
                                                ed['R'+j+'Act'+afterskip[x]][i+pos]=da
                                                # totaldistrib.append(da)
                                            pos=pos+1
                                            thirdPortEndpos.append(pos)
                                        else:
                                            pos=pos+1
                                            thirdPortEndpos.append(pos)
                                    except:
                                        pass
                            else:
                                thirdPortEndpos=secondPortEndpos

                        
                        fourthPortEndpos=[]
                        if port_name[3]=="Warri":
                            fourthPort=Warri_Port
                            pos=TWTvalue["Warri"]+thirdPortEndpos[-1]
                        if port_name[3]=="PH":
                            fourthPort=PH_Port
                            pos=TWTvalue["PH"]+thirdPortEndpos[-1]
                        if port_name[3]=="Lagos":
                            fourthPort=Lagosport
                            pos=TWTvalue["Lagos"]+thirdPortEndpos[-1]
                        if port_name[3]=="Port4":
                            fourthPort=PortFour
                            pos=TWTvalue["Port4"]+thirdPortEndpos[-1]
                        if port_name[3]=="Port5":
                            fourthPort=PortFive
                            pos=TWTvalue["Port5"]+thirdPortEndpos[-1]
                        # pos=lagosEndpos[-1]+TWTvalue["PH"]
                        balance=0
                        for x in range(len(afterskip)):
                            if afterskip[x] in fourthPort and list_[x]>0:
                                da=list_[x]
                                da=da-balance
                                while da>4500:
                                    ed['R'+j+'Act'+afterskip[x]][i+pos]=4500
                                    totaldistrib.append(4500)
                                    da=da-4500
                                    pos=pos+1
                                    fourthPortEndpos.append(pos)
                                if da<4500 and da>=0:
                                    ed['R'+j+'Act'+afterskip[x]][i+pos]=da
                                    totaldistrib.append(da)
                                    balance=4500-da
                                    try:
                                        if list_[x+1]!=0:
                                            if afterskip[x+1] in fourthPort:
                                                if list_[x+1]>balance:
                                                    ed['R'+j+'Act'+afterskip[x+1]][i+pos]=balance
                                                    totaldistrib.append(balance)
                                                else:
                                                    ed['R'+j+'Act'+afterskip[x+1]][i+pos]=list_[x+1]
                                                    totaldistrib.append(list_[x+1])
                                                    balance=balance-list_[x+1]
                                            else:
                                                ed['R'+j+'Act'+afterskip[x]][i+pos]=da
                                                # totaldistrib.append(da)
                                            pos=pos+1
                                            fourthPortEndpos.append(pos)
                                        else:
                                            pos=pos+1
                                            fourthPortEndpos.append(pos)
                                    except:
                                        pass
                            else:
                                fourthPortEndpos=thirdPortEndpos

                        
                        
                        fifthPortEndpos=[]
                        if port_name[4]=="Warri":
                            fifthPort=Warri_Port
                            pos=TWTvalue["Warri"]+fourthPortEndpos[-1]
                        if port_name[4]=="PH":
                            fifthPort=PH_Port
                            pos=TWTvalue["PH"]+fourthPortEndpos[-1]
                        if port_name[4]=="Lagos":
                            fifthPort=Lagosport
                            pos=TWTvalue["Lagos"]+fourthPortEndpos[-1]
                        if port_name[4]=="Port4":
                            fifthPort=PortFour
                            pos=TWTvalue["Port4"]+fourthPortEndpos[-1]
                        if port_name[4]=="Port5":
                            fifthPort=PortFive
                            pos=TWTvalue["Port5"]+fourthPortEndpos[-1]
                        # pos=lagosEndpos[-1]+TWTvalue["PH"]
                        balance=0
                        count=0
                        for x in range(len(afterskip)):
                            if afterskip[x] in fifthPort and list_[x]>0:
                                # if count==0:
                                #     da=totalbalance
                                # else:
                                #     da=list_[x]
                                # count=count+1
                                da=list_[x]
                                da=da-balance
                                while da>4500:
                                    ed['R'+j+'Act'+afterskip[x]][i+pos]=4500
                                    totaldistrib.append(4500)
                                    da=da-4500
                                    pos=pos+1
                                    fifthPortEndpos.append(pos)
                                if da<4500 and da>=0:
                                    ed['R'+j+'Act'+afterskip[x]][i+pos]=da
                                    totaldistrib.append(da)
                                    balance=4500-da
                                    try:
                                        if list_[x+1]!=0:
                                            if afterskip[x+1] in fifthPort:
                                                if list_[x+1]>balance:
                                                    ed['R'+j+'Act'+afterskip[x+1]][i+pos]=balance
                                                    totaldistrib.append(balance)
                                                else:
                                                    ed['R'+j+'Act'+afterskip[x+1]][i+pos]=list_[x+1]
                                                    totaldistrib.append(list_[x+1])
                                                    balance=balance-list_[x+1]
                                            else:
                                                ed['R'+j+'Act'+afterskip[x]][i+pos]=da
                                                # totaldistrib.append(da)
                                            pos=pos+1
                                            fifthPortEndpos.append(pos)
                                        else:
                                            pos=pos+1
                                            fifthPortEndpos.append(pos)
                                    except:
                                        pass
                            else:
                                fifthPortEndpos=fourthPortEndpos
                            


                        print("final total distibuted is here********************************",(totaldistrib))
                        print("sum is here********************************",sum(totaldistrib))
                        totalbalance=order_quantity-sum(totaldistrib)
                        print("final total balance is here********************************",(totalbalance))
                        print('list_---- Once ',list_)


    #                     ed['R'+j+'Act'+k][i] = ed['R'+j+k][i:t].fillna(0).sum()

                        ed['R'+j+'ActFlag'][t+1] =  def_inventory_g
                        m = t + 1
                        while m <= len(ed):
                            print('m',m)
                            #[m+1:]

                            a = checkNegative1(list_ag,m,def_inventory_g)

    #                         a ,list_ = checkNegative1(ed,j,list_ag,m,def_inventory_g)

                            print('a-----',a)
                            if a == None :
                                m = len(ed) + 1
                            else : 
#                                 list_ = contriPlant(ed,j,m,a+1)
#                                 for x in range(len(plants)):
#                                     ed['R'+j+'Act'+plants[x]][m] = list_[x] 
                                ed['R'+j+'ActFlag'][a+1] = def_inventory_g
                                m = a + 1

                        temp = 'Found'
                        #print('break3-----------')
#                         print('**********','R'+j+'Flag',len(ed['R'+j+'Flag'][i+1:len(ed)]),'R'+j+'ActFlag',len(ed['R'+j+'ActFlag'][i+1:len(ed)]))
                        ed['R'+j+'Flag'][i+1:len(ed)] = ed['R'+j+'ActFlag'][i+1:len(ed)]
                        temp1 = ed['R'+j+'ActFlag'][0:i+1] 
                        ed['R'+j+'ActFlag'] = np.NAN
                        ed['R'+j+'ActFlag'] = temp1
                        break

            if temp:
                #print('break2')
                break
    #     if temp:
    #         #print('break2')
    #         break

    return ed


############################## Importing Excel 
##############################


ex = pd.ExcelFile(file_path + '/Inventory optimization working.xlsx')
ed = ex.parse('Sheet2')


para = ex.parse('Parameters')

discharge = [i for i in para.iloc[2,1:].values]
thresh = [i for i in para.iloc[3,1:].values]
order_quantity = [i for i in para.iloc[4:,1].values]
order_quantity = order_quantity[0]
port_name=(list(para["Portname"]))
wt=(list(para["TWT"]))

TWTvalue=dict(zip(port_name,wt))
# print(TWTvalue)

randomplants=[i for i in para.iloc[1,1:]]
plants = ['P1','P2','P3','P4','P5','P6','P7','P8','P9','P10']
types = ['C','L','S']
table_type = ['R','OH','DOI']

Lagosport=['P1','P2','P3']
PH_Port=["P4"]
Warri_Port=["P5"]
PortFour=['P6','P7','P8']
PortFive=['P9','P10']



afterskip=[]
for x in randomplants:
    if x in plants:
        afterskip.append(x)
# print("after skip plant************************",afterskip)


RCP = ed[['RCP1', 'RCP2', 'RCP3', 'RCP4', 'RCP5', 'RCP6', 'RCP7', 'RCP8', 'RCP9','RCP10']]
RLP = ed[['RLP1', 'RLP2', 'RLP3', 'RLP4', 'RLP5', 'RLP6', 'RLP7', 'RLP8', 'RLP9','RLP10']]
RSP = ed[['RSP1', 'RSP2', 'RSP3', 'RSP4', 'RSP5', 'RSP6', 'RSP7', 'RSP8', 'RSP9','RSP10']]
ed.columns[2:]



DOICP = ed[['id','OHCP1', 'OHCP2', 'OHCP3', 'OHCP4', 'OHCP5','OHCP6', 'OHCP7', 'OHCP8', 'OHCP9', 'OHCP10']]
DOILP = ed[['id','OHLP1', 'OHLP2', 'OHLP3','OHLP4', 'OHLP5', 'OHLP6', 'OHLP7', 'OHLP8', 'OHLP9', 'OHLP10']]
DOISP = ed[['id','OHSP1', 'OHSP2', 'OHSP3', 'OHSP4', 'OHSP5', 'OHSP6', 'OHSP7', 'OHSP8', 'OHSP9','OHSP10']]

## Calculating Aggregation
ed['RCAgg']=calculateRCAgg()
ed['RLAgg']=calculateRLAgg()
ed['RSAgg']=calculateRSAgg()
# ed['RCAgg'] = ed['RCP1'].fillna(0) +ed['RCP2'].fillna(0)+ed['RCP3'].fillna(0)+ed['RCP4'].fillna(0)+ed['RCP5'].fillna(0)+ed['RCP6'].fillna(0)+ed['RCP7'].fillna(0)+ed['RCP8'].fillna(0)+ed['RCP9'].fillna(0)+ed['RCP10'].fillna(0)
# ed['RLAgg'] = ed['RLP1'].fillna(0) +ed['RLP2'].fillna(0)+ed['RLP3'].fillna(0)+ed['RLP4'].fillna(0)+ed['RLP5'].fillna(0)+ed['RLP6'].fillna(0)+ed['RLP7'].fillna(0)+ed['RLP8'].fillna(0)+ed['RLP9'].fillna(0)+ed['RLP10'].fillna(0)
# ed['RSAgg'] = ed['RSP1'].fillna(0) +ed['RSP2'].fillna(0)+ed['RSP3'].fillna(0)+ed['RSP4'].fillna(0)+ed['RSP5'].fillna(0)+ed['RSP6'].fillna(0)+ed['RSP7'].fillna(0)+ed['RSP8'].fillna(0)+ed['RSP9'].fillna(0)+ed['RSP10'].fillna(0)



## Actual filling after user inputs value
ed = actualReceived(ed)

##updating output to excel
updatedftoexcel(ed,'Sheet1')


# ############################# integrated
wheattype=['Canadian','SRW - US','LPRussian','LPGerman','LPFrench','LPArgentinaian']
planttype=["TCA","TCB","BL","PH","Warri","Calabar","Apapa","Ikorodu","Ilorin","Kano"]


RCAct=["RCActP1","RCActP2","RCActP3","RCActP4","RCActP5","RCActP6","RCActP7","RCActP8","RCActP9","RCActP10"]
RLAct=["RLActP1","RLActP2","RLActP3","RLActP4","RLActP5","RLActP6","RLActP7","RLActP8","RLActP9","RLActP10"]
RSAct=["RSActP1","RSActP2","RSActP3","RSActP4","RSActP5","RSActP6","RSActP7","RSActP8","RSActP9","RSActP10"]
RActualF=["RCActFlag","RLActFlag","RSActFlag"]
RLPType=["RLTypeP1","RLTypeP2","RLTypeP3","RLTypeP4","RLTypeP5","RLTypeP6","RLTypeP7","RLTypeP8","RLTypeP9","RLTypeP10"]


Lptype=["Rus","Ger","Fre","Arg"]

def Integrated(edgrist):
    Crepllist=[]
    Srepllist=[]
    RusLrepllist=[]
    GerLrepllist=[]
    FreLrepllist=[]
    ArgLrepllist=[]

    for i in range(len(ed)):
        currentDT = datetime.datetime.now()
        d1=(currentDT.strftime("%Y-%m-%d 00:00:00"))
        # print('ed_date ',ed["Day"][i+1],type(ed["Day"][i+1]))
        # print('d1',d1,type(d1))
        if str(ed["Day"][i])==d1:
            for j in types:
                if j=="C":
                    for r in RCAct:
                        # print("**********",ed[r][i])
                        for k in range(i,len(ed)):
                            if ed[r][k]>0:
                                # print(r,k-i)
                                crepl=k-i
                                Crepllist.append(crepl)
                                break
                        else:
                            Crepllist.append(9999)    
                if j=="S":
                    for r in RSAct:
                        # print("**********",ed[r][i])
                        for k in range(i,len(ed)):
                            if ed[r][k]>0:
                                # print(r,k-i)
                                srepl=k-i
                                Srepllist.append(srepl)
                                break
                        else:
                            Srepllist.append(9999)
                
                
                if j=="L":
                    for r in RLPType:
                        # print("**********",ed[r][i])
                        for k in range(i,len(ed)):
                            if ed[r][k]=="Rus" or ed[r][k]=="Russian" or ed[r][k]=="rus":
                                Rlrepl=k-i
                                # print(Rlrepl)
                                RusLrepllist.append(Rlrepl)
                                break
                        else:
                            RusLrepllist.append(9999)
                        
                        for k in range(i,len(ed)):
                            if ed[r][k]=="Ger" or ed[r][k]=="German" or ed[r][k]=="ger":
                                Glrepl=k-i
                                # print(Rlrepl)
                                GerLrepllist.append(Glrepl)
                                break
                        else:
                            GerLrepllist.append(9999)
                        
                        for k in range(i,len(ed)):
                            if ed[r][k]=="Fre" or ed[r][k]=="French" or ed[r][k]=="fre":
                                Flrepl=k-i
                                # print(Rlrepl)
                                FreLrepllist.append(Flrepl)
                                break
                        else:
                            FreLrepllist.append(9999)

                        
                        for k in range(i,len(ed)):
                            if ed[r][k]=="Arg" or ed[r][k]=="Argentinaian" or ed[r][k]=="arg" :
                                Alrepl=k-i
                                # print(Rlrepl)
                                ArgLrepllist.append(Alrepl)
                                break
                        else:
                            ArgLrepllist.append(9999)
                                    
                    
    if len(Crepllist) == 10:
        edgrist["Canadian"]= Crepllist 
    if len(Srepllist) == 10:
        edgrist["SRW - US"]= Srepllist
    if len(RusLrepllist) == 10:
        edgrist["Russian"]= RusLrepllist
    if len(GerLrepllist) == 10:
        edgrist["German"]= GerLrepllist
    if len(FreLrepllist) == 10:
        edgrist["French"]= FreLrepllist
    if len(ArgLrepllist) == 10:
        edgrist["Argentinaian"]= ArgLrepllist
    

    # print(edgrist)
    print("done")
    return edgrist
           
        
# Integrated(edgrist)

exgrist = pd.ExcelFile(file_path + '/Grist Optimizer changed.xlsx')
edgrist = exgrist.parse('Sheet3')

# Actual filling after user inputs value
edgrist = Integrated(edgrist)

##updating output to excel
updatedftoexcel1(edgrist,'Repl')