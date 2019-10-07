import pandas as pd
import numpy as np
import math
import os

file_path = os.path.dirname(os.path.abspath( __file__ ))


############################## Writing in excel 
##############################

def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False, 
                       **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    Parameters:
      filename : File path or existing ExcelWriter
                 (Example: '/path/to/file.xlsx')
      df : dataframe to save to workbook
      sheet_name : Name of sheet which will contain DataFrame.
                   (default: 'Sheet1')
      startrow : upper left cell row to dump data frame.
                 Per default (startrow=None) calculate the last row
                 in the existing DF and write to the next row...
      truncate_sheet : truncate (remove and recreate) [sheet_name]
                       before writing DataFrame to Excel file
      to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                        [can be dictionary]

    Returns: None
    """
    from openpyxl import load_workbook

    import pandas as pd

    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl')

    # Python 2.x: define [FileNotFoundError] exception if it doesn't exist 
    try:
        FileNotFoundError
    except NameError:
        FileNotFoundError = IOError


    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)

        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)

        # copy existing sheets
        writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()
    
    
def updateValExcel(sheet_name,df,index_set,index_r,col,value):
    df.set_index(index_set, inplace=True)
    df[col][index_r] = value
    df.reset_index(level=0, inplace=True)
    updatedftoexcel(df,sheet_name)
    
def updatedftoexcel(df,sheetName):
    append_df_to_excel(file_path + '/Inventory optimization working.xlsx', df, sheet_name=sheetName,startrow=0,index=False)   



excmn = pd.ExcelFile(file_path + '/commonsheet.xlsx')
edcmn = excmn.parse('Sheet1')


############################## Required Suggestion
##############################

def checkNegative(list_ag,loc,def_inventory):
    list_ag = [0 if math.isnan(x) else x for x in list_ag]
#     print(len(list_ag),len(list_ag)-len(list_ag[loc:]))
    list_ag = list_ag[loc:]
    return_loc = 0
    for i in range(len(list_ag)):
        if def_inventory - list_ag[i] >= 0 :
            def_inventory = def_inventory -list_ag[i]
            return_loc = i
#             print(i,list_ag[i])
        else :
#             print(i,def_inventory -list_ag[i])
            break 
    if return_loc > 0:
        return return_loc + loc
    



## Aggregation logic Without Balance

def flagCal(ed):
    ## order quantity 48000
    def_inventory_g = order_quantity
    
    for j in types:
        temp = None
        for i in range(len(ed)):
            for k in plants:
                if ed['DOI'+j+k][i] <= thresh[int(k[1])] :
                    if j == 'C':
                        temp1 = ed['R'+j+'Flag'][0:i] 
                        ed['R'+j+'Flag'] = np.NAN
                        ed['R'+j+'Flag'] = temp1

                        print('Typej=',j,' Plantk=',k,' Rowi=',i)                 
                        ed['R'+j+'Flag'][i] = def_inventory_g 
                        m = i
                        while m <= len(ed) :
                            print('m',m)
                            list_ag = ed['R'+j+'Agg']#[m+1:]
                            a = checkNegative(list_ag,m,def_inventory_g)
                            print('a-----',a)
                            if a == None :
                                m = len(ed) + 1
                            else :  
                                ed['R'+j+'Flag'][a+1] = def_inventory_g
                                m = a + 1

                        temp = 'Found'
                        #print('break3-----------')
                        break
                    if j == 'L':
                        temp1 = ed['R'+j+'Flag'][0:i] 
                        ed['R'+j+'Flag'] = np.NAN
                        ed['R'+j+'Flag'] = temp1

                        print('Typej=',j,' Plantk=',k,' Rowi=',i)                 
                        ed['R'+j+'Flag'][i] = def_inventory_g 
                        m = i
                        while m <= len(ed) :
                            print('m',m)
                            list_ag = ed['R'+j+'Agg']#[m+1:]
                            a = checkNegative(list_ag,m,def_inventory_g)
                            print('a-----',a)
                            if a == None :
                                m = len(ed) + 1
                            else :  
                                ed['R'+j+'Flag'][a+1] = def_inventory_g
                                m = a + 1
                        #flagCal(RLP,i,j,k)
                        temp = 'Found'
                        print('break3-----------')
                        break
                    if j == 'S':
                        temp1 = ed['R'+j+'Flag'][0:i] 
                        ed['R'+j+'Flag'] = np.NAN
                        ed['R'+j+'Flag'] = temp1

                        print('Typej=',j,' Plantk=',k,' Rowi=',i)                 
                        ed['R'+j+'Flag'][i] = def_inventory_g 
                        m = i
                        while m <= len(ed) :
                            print('m',m)
                            list_ag = ed['R'+j+'Agg']#[m+1:]
                            a = checkNegative(list_ag,m,def_inventory_g)
                            print('a-----',a)
                            if a == None :
                                m = len(ed) + 1
                            else :  
                                ed['R'+j+'Flag'][a+1] = def_inventory_g
                                m = a + 1
                        #flagCal(RSP,i,j,k)
                        temp = 'Found'
                        print('break3-----------')
                        break

            if temp:
                #print('break2')
                break
    #     if temp:
    #         #print('break2')
    #         break
    return ed




############################## Importing Excel 
##############################


ex = pd.ExcelFile(file_path + '/Inventory optimization working.xlsx')
ed = ex.parse('Sheet2')

para = ex.parse('Parameters')

discharge = [i for i in para.iloc[2,1:].values]
thresh = [i for i in para.iloc[3,1:].values]
order_quantity = [i for i in para.iloc[4:,1].values]
order_quantity = order_quantity[0]


plants = ['P1','P2','P3','P4','P5','P6','P7','P8','P9','P10']
types = ['C','L','S']
table_type = ['R','OH','DOI']

RCP = ed[['RCP1', 'RCP2', 'RCP3', 'RCP4', 'RCP5', 'RCP6', 'RCP7', 'RCP8', 'RCP9','RCP10']]
RLP = ed[['RLP1', 'RLP2', 'RLP3', 'RLP4', 'RLP5', 'RLP6', 'RLP7', 'RLP8', 'RLP9','RLP10']]
RSP = ed[['RSP1', 'RSP2', 'RSP3', 'RSP4', 'RSP5', 'RSP6', 'RSP7', 'RSP8', 'RSP9','RSP10']]
ed.columns[2:]



DOICP = ed[['id','OHCP1', 'OHCP2', 'OHCP3', 'OHCP4', 'OHCP5','OHCP6', 'OHCP7', 'OHCP8', 'OHCP9', 'OHCP10']]
DOILP = ed[['id','OHLP1', 'OHLP2', 'OHLP3','OHLP4', 'OHLP5', 'OHLP6', 'OHLP7', 'OHLP8', 'OHLP9', 'OHLP10']]
DOISP = ed[['id','OHSP1', 'OHSP2', 'OHSP3', 'OHSP4', 'OHSP5', 'OHSP6', 'OHSP7', 'OHSP8', 'OHSP9','OHSP10']]

## Calculating Aggregation
ed['RCAgg'] = ed['RCP1'].fillna(0) +ed['RCP2'].fillna(0)+ed['RCP3'].fillna(0)+ed['RCP4'].fillna(0)+ed['RCP5'].fillna(0)+ed['RCP6'].fillna(0)+ed['RCP7'].fillna(0)+ed['RCP8'].fillna(0)+ed['RCP9'].fillna(0)+ed['RCP10'].fillna(0)
ed['RLAgg'] = ed['RLP1'].fillna(0) +ed['RLP2'].fillna(0)+ed['RLP3'].fillna(0)+ed['RLP4'].fillna(0)+ed['RLP5'].fillna(0)+ed['RLP6'].fillna(0)+ed['RLP7'].fillna(0)+ed['RLP8'].fillna(0)+ed['RLP9'].fillna(0)+ed['RLP10'].fillna(0)
ed['RSAgg'] = ed['RSP1'].fillna(0) +ed['RSP2'].fillna(0)+ed['RSP3'].fillna(0)+ed['RSP4'].fillna(0)+ed['RSP5'].fillna(0)+ed['RSP6'].fillna(0)+ed['RSP7'].fillna(0)+ed['RSP8'].fillna(0)+ed['RSP9'].fillna(0)+ed['RSP10'].fillna(0)



## Fillin requirement with 48000 fixed value
ed = flagCal(ed)

##updating output to excel
updatedftoexcel(ed,'Sheet1')